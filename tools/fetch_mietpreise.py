#!/usr/bin/env python3
"""
fetch_mietpreise.py — holt den durchschnittlichen Mietzins (Netto-Kaltmiete) pro
Kanton und Zimmerzahl vom Bundesamt für Statistik (BFS) und schreibt
data/mietpreise.json.

Quelle: BFS — "Durchschnittlicher Mietpreis in Franken nach Zimmerzahl und Kanton"
  Asset: je-d-09.03.03.01 (XLSX), via opendata.swiss / dam-api.bfs.admin.ch
  Datenbasis: Strukturerhebung / Volkszählung (Nettomiete, ohne Neben-/Heizkosten).

Verwendung im Tool: Reality-Anchor für den Miet-Input der Earn-Card (Rent-to-Rent).
  WICHTIG — Granularität (nicht überinterpretieren): Das sind KANTONS-Durchschnitte.
  Resort-/City-Mikrolagen weichen stark ab (z.B. Zermatt >> Wallis-Schnitt). Daher nur
  als grober Plausibilitäts-Anker (🟡 MOD), nicht als marktgenauer Wert. Tier sichtbar.

Aktualität: BFS aktualisiert jährlich. Refresh monatlich via refresh-data.yml.
  Bei Fehler bleibt die letzte Snapshot-Datei intakt, _health.json zeigt den Status.

Risiko-Profil:
- dam-api.bfs.admin.ch ist die offizielle BFS-Asset-Infrastruktur (stabil)
- XLSX-Layout: Kantone als Zeilen, Zimmerzahl-Spalten (Total,1..6) je mit
  Vertrauensintervall. Spalten per Header-Zeile ('Total') aufgelöst → robust.
- Abhängigkeit: openpyxl (in refresh-data.yml via pip)
- Fallback: bei Fehler alte mietpreise.json behalten + _health.json error
"""

import io
import json
import os
import sys
import time
import urllib.request
import zipfile
from datetime import datetime, timezone

# BFS-Asset je-d-09.03.03.01 (deutsche Version, XLSX)
XLS_URL = "https://dam-api.bfs.admin.ch/hub/api/dam/assets/36398431/master"
SOURCE_URL = "https://www.bfs.admin.ch/asset/de/je-d-09.03.03.01"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "data"))
OUT_FILE = os.path.join(DATA_DIR, "mietpreise.json")
HEALTH_FILE = os.path.join(DATA_DIR, "_health.json")
USER_AGENT = "SwissSTR-Intelligence/0.9 (https://github.com/Adrianmaagg/swissstr) Python/urllib"

# BFS-Kantonsname (deutsch) → 2-Buchstaben-Code (wie in js/data.js)
CANTON_CODE = {
    "Zürich": "ZH", "Bern": "BE", "Luzern": "LU", "Uri": "UR", "Schwyz": "SZ",
    "Obwalden": "OW", "Nidwalden": "NW", "Glarus": "GL", "Zug": "ZG",
    "Freiburg": "FR", "Solothurn": "SO", "Basel-Stadt": "BS",
    "Basel-Landschaft": "BL", "Schaffhausen": "SH", "Appenzell A.Rh.": "AR",
    "Appenzell I.Rh.": "AI", "St.Gallen": "SG", "St. Gallen": "SG",
    "Graubünden": "GR", "Aargau": "AG", "Thurgau": "TG", "Tessin": "TI",
    "Waadt": "VD", "Wallis": "VS", "Neuenburg": "NE", "Genf": "GE", "Jura": "JU",
}
ROOM_LABELS = ["Total", "1", "2", "3", "4", "5", "6"]


def _get(url, raw=True, timeout=120):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        data = r.read()
    return data if raw else json.loads(data.decode("utf-8"))


def num(v):
    """Zahl oder None ('X'/'*'/Text → None)."""
    if isinstance(v, (int, float)):
        return round(float(v))
    return None


def parse_xls(blob):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    # Neuestes Jahr-Sheet (Sheet-Namen sind Jahreszahlen)
    year_sheets = [s for s in wb.sheetnames if s.strip().isdigit()]
    if not year_sheets:
        raise RuntimeError("XLSX: keine Jahr-Sheets gefunden")
    year = max(year_sheets, key=lambda s: int(s))
    ws = wb[year]
    rows = list(ws.iter_rows(values_only=True))

    # Header-Zeile finden (enthält 'Total' + Raumzahlen)
    header_idx = None
    for i, row in enumerate(rows[:10]):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if "Total" in vals and "1" in vals and "4" in vals:
            header_idx = i
            break
    if header_idx is None:
        raise RuntimeError("XLSX: Header-Zeile mit 'Total'/Raumzahlen nicht gefunden")
    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    # Spalten-Index je Raum-Label (Miete steht in der Label-Spalte, CI in der nächsten)
    col_of = {}
    for label in ROOM_LABELS:
        if label in header:
            col_of[label] = header.index(label)

    cantons = {}
    schweiz = None
    for row in rows[header_idx + 1:]:
        name = (str(row[0]).strip() if row and row[0] is not None else "")
        if not name:
            continue
        rec = {
            "total": num(row[col_of["Total"]]) if "Total" in col_of else None,
            "rooms": {lbl: num(row[col_of[lbl]]) for lbl in ROOM_LABELS[1:] if lbl in col_of},
        }
        if name == "Schweiz":
            schweiz = rec
        elif name in CANTON_CODE:
            cantons[CANTON_CODE[name]] = rec
    return year, schweiz, cantons


def build_payload(year, schweiz, cantons):
    return {
        "_meta": {
            "source": "BFS — Durchschnittlicher Mietpreis in Franken nach Zimmerzahl und Kanton (je-d-09.03.03.01)",
            "source_url": SOURCE_URL,
            "year": year,
            "license": "BFS / opendata.swiss — frei nutzbar",
            "fetched_at": datetime.now(timezone.utc).isoformat(),
            "count_cantons": len(cantons),
            "unit": "CHF/Monat Nettomiete (ohne Neben-/Heizkosten)",
            "note": (
                "Durchschnittliche Netto-Kaltmiete pro Kanton und Zimmerzahl (BFS "
                "Strukturerhebung). KANTONS-Durchschnitt — Resort-/City-Mikrolagen weichen "
                "stark ab. Im Tool nur als grober Reality-Anchor (🟡 MOD) für den Miet-Input, "
                "nicht als marktgenauer Wert."
            ),
        },
        "schweiz": schweiz,
        "cantons": cantons,
    }


def update_health(status, year=None, count=None, error=None):
    if not os.path.exists(HEALTH_FILE):
        print("WARN: _health.json missing, skipping health update")
        return
    with open(HEALTH_FILE, "r", encoding="utf-8") as f:
        health = json.load(f)
    health.setdefault("sources", {})
    entry = health["sources"].setdefault("bfs_mietpreise", {
        "name": "BFS Durchschnittsmiete (Kanton × Zimmerzahl)",
        "url": SOURCE_URL,
        "snapshot_file": "data/mietpreise.json",
        "expected_frequency": "yearly",
        "note": "BFS Strukturerhebung, jährlich. Reality-Anchor für Miet-Input (Kantons-Schnitt).",
    })
    now = datetime.now(timezone.utc).isoformat()
    entry["last_attempt"] = now
    if status == "ok":
        entry["last_success"] = now
        entry["status"] = "ok"
        entry["period_covered"] = str(year)
        entry["covered_cantons"] = count
        entry.pop("error", None)
    else:
        entry["status"] = "error"
        if error:
            entry["error"] = error
    health["generated_at"] = now
    with open(HEALTH_FILE, "w", encoding="utf-8") as f:
        json.dump(health, f, ensure_ascii=False, indent=2)


def main():
    try:
        print("Fetching BFS Durchschnittsmiete (Kanton × Zimmerzahl)...")
        t0 = time.time()
        blob = _get(XLS_URL)
        print(f"Downloaded {len(blob)} bytes")
        year, schweiz, cantons = parse_xls(blob)
        payload = build_payload(year, schweiz, cantons)
        n = len(cantons)
        if n < 20:
            raise RuntimeError(f"Sanity check failed: only {n} cantons (expected 26)")
        if not schweiz or not schweiz.get("total"):
            raise RuntimeError("Sanity check failed: Schweiz-Total fehlt")
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(OUT_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        update_health("ok", year=year, count=n)
        print(f"OK — {n} cantons, year {year}, CH-Total CHF {schweiz['total']} in {time.time()-t0:.1f}s")
        return 0
    except Exception as e:
        print(f"FAILED: {e}", file=sys.stderr)
        update_health("error", error=str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())
