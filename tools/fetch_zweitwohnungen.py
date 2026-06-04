#!/usr/bin/env python3
"""
fetch_zweitwohnungen.py — holt das ARE-Wohnungsinventar (Zweitwohnungsanteil pro
Schweizer Gemeinde) und schreibt data/zweitwohnungen.json.

Quelle: Bundesamt für Raumentwicklung ARE — "Wohnungsinventar und Zweitwohnungsanteil"
  Layer: ch.are.wohnungsinventar-zweitwohnungsanteil
  Portal: https://opendata.swiss/de/dataset/wohnungsinventar-und-zweitwohnungsanteil
  Download: geo.admin.ch STAC-API (jährliche/halbjährliche Releases als XLSX)

Hintergrund (Zweitwohnungsgesetz ZWG, SR 702):
  Gemeinden mit einem Zweitwohnungsanteil von 20% oder mehr dürfen grundsätzlich keine
  neuen Zweitwohnungen mehr bewilligen. Die ARE bestimmt den Status jährlich per 31.3.

WICHTIG — Status vs. 20%-Schwelle (nicht glätten):
  Die rohe 20%-Schwelle (ZWG_3120 >= 20) und der offizielle ARE-Status weichen leicht ab,
  weil Gemeinden via Sonderverfahren einen tieferen Erstwohnungsanteil nachweisen können.
  Wir speichern BEIDES: den echten Zweitwohnungsanteil (pct) UND den offiziellen
  restricted-Flag (Status==1). Der Status ist die rechtlich massgebende Bestimmung.

Aktualität: ARE publiziert per 31.3. (+ teils 31.10.). Refresh monatlich via
  .github/workflows/refresh-data.yml. Bei API-/Parse-Fehler bleibt die letzte
  Snapshot-Datei intakt, _health.json zeigt den Status.

Risiko-Profil (für Adrian's Risiko-Filter):
- geo.admin.ch STAC-API ist die offizielle Bundes-Geodaten-Infrastruktur (stabil)
- Schema der XLSB-Tabelle: Spalten per Header-Name aufgelöst (Gem_No, ZWG_3120, Status),
  robust gegen Spalten-Umordnung; bricht sichtbar wenn Header fehlen
- Abhängigkeit: openpyxl (in refresh-data.yml via pip installiert)
- Fallback: bei Fehler alte zweitwohnungen.json behalten + _health.json error
"""

import io
import json
import os
import sys
import time
import urllib.request
import zipfile
from datetime import datetime, timezone

STAC_COLLECTION = (
    "https://data.geo.admin.ch/api/stac/v0.9/collections/"
    "ch.are.wohnungsinventar-zweitwohnungsanteil"
)
PORTAL_URL = "https://opendata.swiss/de/dataset/wohnungsinventar-und-zweitwohnungsanteil"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, "..", "data"))
OUT_FILE = os.path.join(DATA_DIR, "zweitwohnungen.json")
HEALTH_FILE = os.path.join(DATA_DIR, "_health.json")
USER_AGENT = "SwissSTR-Intelligence/0.9 (https://github.com/Adrianmaagg/swissstr) Python/urllib"

ZWG_THRESHOLD = 20.0  # ZWG SR 702: ab 20% Zweitwohnungsanteil grundsätzlich Bewilligungs-Stopp


def _get(url, raw=False, timeout=120):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        data = r.read()
    return data if raw else json.loads(data.decode("utf-8"))


def find_latest_xlsx():
    """STAC-Items abrufen, neuestes Release wählen, XLSX-ZIP-Asset zurückgeben."""
    items = _get(STAC_COLLECTION + "/items")
    feats = items.get("features", [])
    if not feats:
        raise RuntimeError("STAC: keine Items gefunden")
    # IDs enden auf Datum (2017, 2019-03, ..., 2026-03) — lexikografischer Sort = chronologisch
    feats.sort(key=lambda f: f.get("id", ""))
    latest = feats[-1]
    release = latest["id"].split("_")[-1]  # z.B. "2026-03"
    href = None
    for asset in (latest.get("assets") or {}).values():
        h = asset.get("href", "")
        if h.endswith(".xlsx.zip"):
            href = h
            break
    if not href:
        raise RuntimeError(f"STAC: kein .xlsx.zip-Asset im Release {release}")
    return release, href


def parse_xlsx(blob):
    """XLSX-ZIP → Liste von Gemeinde-Records. Spalten per Header-Name aufgelöst."""
    import openpyxl  # in refresh-data.yml via pip installiert

    z = zipfile.ZipFile(io.BytesIO(blob))
    xlsx_name = next(n for n in z.namelist() if n.lower().endswith(".xlsx"))
    wb = openpyxl.load_workbook(io.BytesIO(z.read(xlsx_name)), read_only=True, data_only=True)

    # Daten-Sheet = jenes mit einer Header-Zeile, die 'Gem_No' enthält (Sheet-Name ist dynamisch)
    target_ws = None
    header = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and "Gem_No" in row:
                target_ws = ws
                header = list(row)
                break
        if target_ws:
            break
    if not target_ws:
        raise RuntimeError("XLSX: Daten-Sheet mit Header 'Gem_No' nicht gefunden")

    def col(name):
        if name not in header:
            raise RuntimeError(f"XLSX: Spalte '{name}' fehlt — Schema geändert?")
        return header.index(name)

    c_gem, c_name, c_kt = col("Gem_No"), col("Name"), col("Kt_Kz")
    c_total, c_pct = col("ZWG_3150"), col("ZWG_3120")
    c_verf, c_status = col("Verfahren"), col("Status")

    started = False
    records = []
    for row in target_ws.iter_rows(min_row=1, values_only=True):
        if not started:
            if row and "Gem_No" in row:
                started = True
            continue
        gem = row[c_gem]
        if gem is None or not isinstance(gem, (int, float)):
            continue
        pct = row[c_pct]
        status = row[c_status]
        records.append({
            "name": row[c_name],
            "kt": row[c_kt],
            "pct": round(float(pct), 2) if isinstance(pct, (int, float)) else None,
            "total": int(row[c_total]) if isinstance(row[c_total], (int, float)) else None,
            "restricted": (status == 1),   # offizieller ARE-Status (massgebend)
            "verfahren": int(row[c_verf]) if isinstance(row[c_verf], (int, float)) else None,
            "_gem": int(gem),
        })
    return records


def build_payload(release, records):
    communes = {}
    for r in records:
        bfs = str(r.pop("_gem"))
        communes[bfs] = r
    n = len(communes)
    over20 = sum(1 for r in communes.values() if r["pct"] is not None and r["pct"] >= ZWG_THRESHOLD)
    restricted = sum(1 for r in communes.values() if r["restricted"])
    return {
        "_meta": {
            "source": "Bundesamt für Raumentwicklung ARE — Wohnungsinventar und Zweitwohnungsanteil (ch.are.wohnungsinventar-zweitwohnungsanteil)",
            "source_url": PORTAL_URL,
            "release": release,
            "license": "geo.admin.ch / opendata.swiss — frei nutzbar (BGDI)",
            "fetched_at": datetime.now(timezone.utc).isoformat(),
            "count": n,
            "count_over_20pct": over20,
            "count_restricted": restricted,
            "threshold_pct": ZWG_THRESHOLD,
            "note": (
                "Zweitwohnungsanteil pro Gemeinde (ZWG SR 702). 'restricted' = offizieller "
                "ARE-Status (Status==1, rechtlich massgebend): keine neuen Zweitwohnungen "
                "bewilligbar. Weicht leicht von der reinen 20%-Schwelle ab (count_over_20pct), "
                "da Sonderverfahren möglich sind — beide Zahlen bewusst dokumentiert. "
                "Für Rent-to-Rent von BESTANDSwohnungen nicht direkt verbietend, aber starkes "
                "regulatorisches Sensibilitäts-Signal (Proxy)."
            ),
        },
        "communes": communes,
    }


def update_health(status, release=None, count=None, restricted=None, error=None):
    if not os.path.exists(HEALTH_FILE):
        print("WARN: _health.json missing, skipping health update")
        return
    with open(HEALTH_FILE, "r", encoding="utf-8") as f:
        health = json.load(f)
    health.setdefault("sources", {})
    entry = health["sources"].setdefault("are_zweitwohnungen", {
        "name": "ARE Wohnungsinventar / Zweitwohnungsanteil",
        "url": PORTAL_URL,
        "snapshot_file": "data/zweitwohnungen.json",
        "expected_frequency": "yearly",
        "note": "ARE publiziert per 31.3. Genutzt für ZWG-Cap-Regulierungs-Signal pro Markt.",
    })
    now = datetime.now(timezone.utc).isoformat()
    entry["last_attempt"] = now
    if status == "ok":
        entry["last_success"] = now
        entry["status"] = "ok"
        entry["period_covered"] = release
        entry["covered_communes"] = count
        entry["restricted_communes"] = restricted
        entry.pop("error", None)
    else:
        entry["status"] = "error"
        if error:
            entry["error"] = error
    health["generated_at"] = now
    with open(HEALTH_FILE, "w", encoding="utf-8") as f:
        json.dump(health, f, ensure_ascii=False, indent=2)


def main():
    try:
        print("Fetching ARE Wohnungsinventar (Zweitwohnungsanteil)...")
        t0 = time.time()
        release, href = find_latest_xlsx()
        print(f"Latest release: {release}")
        blob = _get(href, raw=True)
        print(f"Downloaded {len(blob)} bytes")
        records = parse_xlsx(blob)
        payload = build_payload(release, records)
        n = payload["_meta"]["count"]
        restricted = payload["_meta"]["count_restricted"]
        # Sanity: ~2100 Gemeinden, ~250-450 restricted
        if n < 2000:
            raise RuntimeError(f"Sanity check failed: only {n} communes (expected ~2100)")
        if not (250 <= restricted <= 450):
            raise RuntimeError(f"Sanity check failed: {restricted} restricted (expected ~250-450)")
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(OUT_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        update_health("ok", release=release, count=n, restricted=restricted)
        print(f"OK — {n} communes, {restricted} restricted (release {release}) in {time.time()-t0:.1f}s")
        return 0
    except Exception as e:
        print(f"FAILED: {e}", file=sys.stderr)
        update_health("error", error=str(e))
        return 1


if __name__ == "__main__":
    sys.exit(main())
